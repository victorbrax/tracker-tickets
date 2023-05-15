# Imports
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import string
from utils.dfstyles import row_style

# Sheetfiles Path
file_path1 = 'sheetfiles\Planilha 1.xlsx'
file_path2 = 'sheetfiles\Planilha 2.csv'
save_path = 'sheetfiles/Tracker Tickets.xlsx'

# Read and drop the junk
df1 = pd.read_excel(file_path1)
df2 = pd.read_csv(file_path2, sep=';')

df1_junk = [0, 2, 3, 5, 6] + list(range(15, 20))
df2_junk = list(range(4, 10)) + [11, 12, 13]

df1.drop(df1.columns[df1_junk], axis=1, inplace=True)
df2.drop(df2.columns[df2_junk], axis=1, inplace=True)

# Merged DF Customs
df_merged = df1.merge(df2, left_on='Our Tickets', right_on='No.', how='left')

df_merged.drop(['No.'], axis=1, inplace=True)

df_merged.rename(columns={'Notes': 'Comentários', 
                          'Title': 'Título',
                          'Our Tickets': 'Tickets',
                          'Date': 'Data de Criação WHD',
                          'Updated': 'Atualização',
                          'Request Detail': 'Descrição'}, inplace=True)

# Apply yellow highlight rows and save
df_merged = df_merged.style.apply(row_style, axis=1)
df_merged.to_excel(save_path , index=False)

# Rename Worksheet
wb = openpyxl.load_workbook(save_path)
ws = wb['Sheet1']
ws.title = 'Tracker Tickets'

# Set up style lists and objects
letters = string.ascii_uppercase
cells = [letter + '1' for letter in letters if letter <= 'N'] # List of cells from 'A1' to 'M1'

# Colors
fill = PatternFill(patternType='solid', fgColor='00008B')
font = Font(bold=True, color='ffffff', size=12)

# Apply
for column in ws.iter_cols():
    for cell in column:
        if cell.coordinate in cells:
            cell.fill = fill
            cell.font = font

# Finish
wb.save(save_path)
wb.close()